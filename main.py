import pandas as pd
import streamlit as st
import os
import plotly.graph_objects as go
from agents.data_analyst_agent import DataAnalystAgent
from structure_data_agent import name_unnamed_features
from agents.agent_chat import ChatAgent # for general chatting when datasets are not yet uploaded. 
from pdf_replicator import replicate_pdf
import utils.utils as utils
import json
import datetime
import time
from secretary_ai_functions import send_email
import traceback 
import warnings
import copy
import math
import openpyxl 
import re
import io


# Define the path for the instructions file
INSTRUCTIONS_FILE = "user_instructions.txt"
FEEDBACK_FILE = "feedback.txt" 
DAILY_LOG_FILEPATH = "message_history/daily_chat_log.json"


def _make_column_names_unique(column_names):
    name_counts = {}
    new_columns = []
    for col in column_names:
        if col is None:
            col_str = "Unnamed"
        elif not isinstance(col, str):
            col_str = str(col).strip() # Good: handles non-strings, strips whitespace
        else:
            col_str = col.strip()      # Good: handles strings, strips whitespace

        if not col_str: 
            col_str = "Unnamed"        # Good: handles empty strings after stripping
        
        # Current logic:
        if col_str not in name_counts:
            name_counts[col_str] = 1
            new_columns.append(col_str) 
        else:
            name_counts[col_str] += 1
            new_columns.append(f"{col_str}_{name_counts[col_str]}")
    return new_columns


def _process_dataframe(df):
    """
    A helper function to apply a standard set of cleaning and transposing rules.
    This version also removes columns that are duplicates by content.
    """
    # 1. Remove rows and columns that are completely empty
    cleaned_df = df.dropna(how='all', axis=0).dropna(how='all', axis=1)

    # 2. Remove rows that are more than 40% empty
    if cleaned_df.empty:
        return cleaned_df
        
    min_non_empty_values = math.ceil(len(cleaned_df.columns) * 0.6)
    cleaned_df = cleaned_df.dropna(thresh=min_non_empty_values, axis=0)
    
    if cleaned_df.empty:
        return cleaned_df

    # 3. Set the first column as the index to prepare for transposing
    feature_col_name = cleaned_df.columns[0]
    feature_col_series = cleaned_df[feature_col_name].fillna('Unnamed Feature') # Ensure index has no NaNs
    indexed_df = cleaned_df.set_index(feature_col_series)
    if feature_col_name in indexed_df.columns: # Drop the original column if it wasn't the only one
         indexed_df = indexed_df.drop(columns=feature_col_name)
    
    # 4. Transpose the DataFrame
    transposed_df = indexed_df.T

    # 5. Make column names unique (handling name collisions)
    if not transposed_df.empty:
        transposed_df.columns = _make_column_names_unique(list(transposed_df.columns))
    
    # --- NEW STEP 6: Remove columns that are duplicates by content ---
    if not transposed_df.empty:
        columns_to_keep = []
        seen_column_signatures = set()

        for col_name in transposed_df.columns:
            current_series = transposed_df[col_name]
            # Create a signature: convert to string, replace 'nan', make tuple
            # This makes NaNs comparable for uniqueness
            series_signature = tuple(current_series.astype(str).str.replace('nan', '_NaN_Placeholder_').values)
            
            if series_signature not in seen_column_signatures:
                columns_to_keep.append(col_name)
                seen_column_signatures.add(series_signature)
            # Else, this column's content is a duplicate of one already kept, so we skip it.
        
        transposed_df = transposed_df[columns_to_keep]

    return transposed_df


def convert_sum_to_addition(formula):
    """Converts simple SUM formulas like '=SUM(A1,B2)' to '=A1+B2'."""
    match = re.match(r'=SUM\(([^:]+)\)', formula, re.IGNORECASE)
    if match and ',' in match.group(1):
        args = [arg.strip() for arg in match.group(1).split(',')]
        return '=' + '+'.join(args)
    return formula


def get_final_data_views(file_source, sheet_name=None):
    """
    Analyzes an Excel sheet and returns two cleaned DataFrames and 
    a dictionary of human-readable VERTICAL equations using unique feature names.
    """
    try:
        if isinstance(file_source, str):
            wb_formulas = openpyxl.load_workbook(file_source, data_only=False)
            wb_values = openpyxl.load_workbook(file_source, data_only=True)
        else:
            file_source.seek(0)
            file_content_bytes = file_source.read()
            wb_formulas = openpyxl.load_workbook(io.BytesIO(file_content_bytes), data_only=False)
            wb_values = openpyxl.load_workbook(io.BytesIO(file_content_bytes), data_only=True)
        
        active_sheet_title = wb_formulas.active.title
        if sheet_name and sheet_name in wb_formulas.sheetnames:
            sheet_formulas = wb_formulas[sheet_name]; sheet_values = wb_values[sheet_name]
        else:
            sheet_formulas = wb_formulas.active; sheet_values = wb_values.active
    except Exception as e:
        st.error(f"Error loading workbook: {e}"); return None, None, None

    # First, load all data into a raw dataframe to determine structure
    initial_data = []
    for row in sheet_values.iter_rows():
        initial_data.append([cell.value for cell in row])
    df_raw = pd.DataFrame(initial_data)

    # --- 1. Pre-process to get unique names and their original row locations ---
    # Apply the same cleaning steps used in _process_dataframe
    temp_df = df_raw.dropna(how='all', axis=0).dropna(how='all', axis=1)
    if not temp_df.empty:
        min_non_empty = math.ceil(len(temp_df.columns) * 0.6)
        temp_df = temp_df.dropna(thresh=min_non_empty, axis=0)

    # Generate the unique, suffixed names from the first column of the cleaned data
    feature_col_series = temp_df[temp_df.columns[0]].fillna('Unnamed Feature')
    unique_names = _make_column_names_unique(list(feature_col_series))
    
    # Map the original Excel row index (from temp_df.index) to its new unique name
    row_index_to_unique_name = dict(zip(temp_df.index, unique_names))

    # --- 2. Build the detailed coordinate-to-UNIQUE-name map ---
    coord_to_unique_name_map = {}
    for row_idx, unique_name in row_index_to_unique_name.items():
        # Map all potential cells in this original row to the unique name
        for col_idx in range(sheet_formulas.max_column):
            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
            coord = f"{col_letter}{row_idx + 1}" # +1 because openpyxl rows are 1-based
            coord_to_unique_name_map[coord] = unique_name

    # --- 3. Build DataFrames and the FINAL Equations Dictionary using the smart map ---
    values_data, hybrid_data, equations_dict = [], [], {}
    for row in sheet_formulas.iter_rows():
        values_row, hybrid_row = [], []
        for cell in row:
            value = sheet_values[cell.coordinate].value
            values_row.append(value)
            hybrid_content = value
            
            if cell.data_type == 'f':
                formula_string = cell.value
                if '!' not in formula_string and '[' not in formula_string:
                    hybrid_content = formula_string
                    
                    is_horizontal, formula_row_num = True, cell.row
                    referenced_rows = [int(r) for r in re.findall(r'[A-Z]+([0-9]+)', formula_string)]
                    if not referenced_rows: is_horizontal = False
                    else:
                        for ref_row in referenced_rows:
                            if ref_row != formula_row_num:
                                is_horizontal = False; break
                    
                    if not is_horizontal:
                        # KEY is the unique name of the cell holding the formula
                        key = coord_to_unique_name_map.get(cell.coordinate)
                        
                        if key and key not in equations_dict:
                            processed_formula = convert_sum_to_addition(formula_string)
                            readable_formula = processed_formula.lstrip('=')
                            
                            cell_refs = re.findall(r'([A-Z]+[0-9]+)', readable_formula)
                            for ref in sorted(list(set(cell_refs)), key=len, reverse=True):
                                # Use the UNIQUE name from our new map
                                mapped_name = coord_to_unique_name_map.get(ref, ref)
                                readable_formula = re.sub(r'\b' + ref + r'\b', mapped_name, readable_formula)

                            readable_formula = re.sub(r'([*\/+\-])', r' \1 ', readable_formula)
                            readable_formula = re.sub(r'\s+', ' ', readable_formula).strip()
                            
                            equations_dict[key] = readable_formula

            hybrid_row.append(hybrid_content)
        values_data.append(values_row)
        hybrid_data.append(hybrid_row)
        
    df_values_raw = pd.DataFrame(values_data)

    final_values_df = _process_dataframe(df_values_raw.copy())

    return final_values_df, equations_dict

def convert_to_float_if_numeric(df: pd.DataFrame) -> pd.DataFrame:
    """
    Goes column by column through a DataFrame and converts the
    data type to float if the entire column is numeric.

    Args:
        df: The input pandas DataFrame.

    Returns:
        The DataFrame with numeric columns converted to float type.
    """
    for col in df.columns:
        # Attempt to convert the column to a numeric type.
        # 'errors=coerce' will turn any non-numeric values into NaN (Not a Number).
        numeric_series = pd.to_numeric(df[col], errors='coerce')

        # If the conversion doesn't result in all values being NaN,
        # it means the column was numeric. Then we can change the type.
        if not numeric_series.isnull().all():
            df[col] = numeric_series.astype(float)
    return df

def load_data(uploaded_file_obj):
    """
    Loads data from an uploaded file.
    Returns a pandas DataFrame for CSVs or a dictionary of DataFrames for Excel files.
    For PDFs, returns the output from replicate_pdf.
    """
    try:
        uploaded_file_obj.seek(0)
        file_name = uploaded_file_obj.name.lower()
        if file_name.endswith('.csv'):
            return pd.read_csv(uploaded_file_obj)
        elif file_name.endswith(('.xls', '.xlsx', '.xlsm', '.xlsb')):
            # Pass the file object directly, openpyxl (used by read_excel) can handle it
            return pd.read_excel(uploaded_file_obj, sheet_name=None, engine='openpyxl')
        elif file_name.endswith('.pdf'):
            uploaded_file_obj.seek(0)
            return replicate_pdf(uploaded_file_obj)
        else:
            st.error(f"Unsupported file type: {uploaded_file_obj.name}")
            return None
    except Exception as e:
        st.error(f'Error reading file {uploaded_file_obj.name}: {e}')
        return None

def upload_and_format_files():
    """
    Handles file uploads, standardizes data, capitalizes file and sheet names, 
    and processes unstructured sheets using a cache to avoid re-computation.
    """
    uploaded_files = st.file_uploader(
        'Upload Data Files',
        type=['csv', 'xlsx', 'xls', 'xlsm', 'xlsb', 'pdf'],
        accept_multiple_files=True,
        key="file_uploader_capitalized_integrated"
    )
    
    dataframes_dict = {}
    
    # <<< MODIFIED >>> This will now collect equations from all processed files
    all_equations = {}

    is_structured = True

    if uploaded_files:
        for uploaded_file_obj in uploaded_files:
            original_file_name = uploaded_file_obj.name
            capitalized_file_name = ""
            try:
                uploaded_file_obj.seek(0)
                base_name, extension = os.path.splitext(original_file_name)
                capitalized_file_name = base_name.upper() + extension.upper()

                loaded_data = load_data(uploaded_file_obj) 

                if loaded_data is None:
                    continue

                standardized_data_output = None

                # Updated check for PDF output (now a dict with integer keys for pages)
                if isinstance(loaded_data, dict) and all(isinstance(key, int) for key in loaded_data.keys()):
                    dataframes_dict[capitalized_file_name] = loaded_data
                    is_structured = False
                    # placing the full pdf file in st.session_state
                    st.session_state["pdf_file_objects"][capitalized_file_name] = uploaded_file_obj
                    continue

                if isinstance(loaded_data, dict): # Excel file (assumes string keys for sheets)
                    capitalized_standardized_sheets = {}
                    for original_sheet_name, df_sheet_original in loaded_data.items():
                        capitalized_sheet_name = original_sheet_name.upper()
                        df_for_processing = df_sheet_original

                        if isinstance(df_for_processing, pd.DataFrame):
                            is_structured = utils.is_dataframe_structured(df_for_processing)
                            
                            if not is_structured:
                                # <<< --- NEW CACHING LOGIC START --- >>>
                                
                                # Create a unique key for this specific sheet
                                cache_key = f"{capitalized_file_name}/{capitalized_sheet_name}"

                                # 1. CHECK THE CACHE FIRST
                                if cache_key in st.session_state.processed_sheets_cache:
                                    # If found, use the cached result and skip the slow processing
                                    st.write(f"Loading '{capitalized_sheet_name}' from cache...") # Optional: for debugging
                                    cached_data = st.session_state.processed_sheets_cache[cache_key]
                                    df_for_processing = cached_data['dataframe']
                                    equations_dict = cached_data['equations']
                                    if equations_dict:
                                        all_equations.update(equations_dict)

                                else:
                                    # 2. IF NOT IN CACHE, DO THE HEAVY LIFTING ONCE
                                    with st.spinner(f"Performing one-time processing on sheet: {capitalized_sheet_name}..."):
                                        uploaded_file_obj.seek(0)
                                        processed_values_df, equations_dict = get_final_data_views(
                                            uploaded_file_obj, 
                                            sheet_name=original_sheet_name
                                        ) 

                                    if processed_values_df is not None and not processed_values_df.empty:
                                        # Run your subsequent processing steps
                                        df_for_processing = name_unnamed_features(processed_values_df)
                                        df_for_processing = convert_to_float_if_numeric(df_for_processing)

                                        # 3. SAVE THE NEW RESULT TO THE CACHE
                                        st.session_state.processed_sheets_cache[cache_key] = {
                                            'dataframe': df_for_processing,
                                            'equations': equations_dict
                                        }
                                        if equations_dict:
                                            all_equations.update(equations_dict)
                                    else:
                                        st.warning(f"Could not structure sheet '{capitalized_sheet_name}'. Using original format.")
                                        # df_for_processing remains df_sheet_original

                                # <<< --- NEW CACHING LOGIC END --- >>>

                            # Standardize the chosen DataFrame (either original, cached, or newly processed)
                            capitalized_standardized_sheets[capitalized_sheet_name] = utils.standardize_file(df_for_processing)
                        else:
                            capitalized_standardized_sheets[capitalized_sheet_name] = df_for_processing
                            st.warning(f"Sheet '{capitalized_sheet_name}' in '{capitalized_file_name}' is not a DataFrame.")
                    
                    standardized_data_output = capitalized_standardized_sheets
                
                elif isinstance(loaded_data, pd.DataFrame): # CSV file
                    standardized_data_output = utils.standardize_file(loaded_data)
                else:
                    st.warning(f"Unexpected data type after loading {capitalized_file_name}: {type(loaded_data)}")
                    continue

                if standardized_data_output is not None:
                    dataframes_dict[capitalized_file_name] = standardized_data_output
            except Exception as e:
                error_trace = traceback.format_exc()
                error_msg = f"Failed to process file {capitalized_file_name or original_file_name}: {e}\nTraceback:\n{error_trace}"
                st.error(error_msg)
                print(error_msg)

    if not is_structured:
        return dataframes_dict, all_equations
    else:
        return dataframes_dict, None

def load_instructions(filepath=INSTRUCTIONS_FILE):
    """Loads instructions from the specified file."""
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r') as f:
                return f.read()
        except Exception as e:
            st.error(f"Error loading instructions: {e}")
            return "None"
    return "None"

def save_instructions(instructions, filepath=INSTRUCTIONS_FILE):
    """Saves instructions to the specified file."""
    try:
        with open(filepath, 'w') as f:
            f.write(instructions)
        return True
    except Exception as e:
        st.error(f"Error saving instructions: {e}")
        return False



def display_chat_history():
    """
    Displays the chat history stored in session state and populates
    st.session_state.message_history with a simplified version.
    """
    if 'messages' not in st.session_state or not st.session_state.messages:
        return

    # Ensure message_history is initialized as a dictionary
    if 'message_history' not in st.session_state or not isinstance(st.session_state.message_history, dict):
        st.session_state.message_history = {}

    # 1. Iterate through prompt_ids from st.session_state.messages
    for prompt_id, messages_list in st.session_state.messages.items():
        # For each prompt_id, we will build a list of simplified history messages
        current_prompt_history_list = []
        
        is_first_user_message_for_display = True # For Streamlit display logic

        for message in messages_list: # Iterate through each message object in the list for the prompt_id
            role = message.get('role', 'unknown')
            content = message.get('content')
            
            # This list will store string parts that make up the simplified content for the current message
            history_parts_for_this_message = []

            # --- User Message Handling ---
            if role == 'user':
                # Display logic (only the first user message in a thread is explicitly written with st.chat_message by original logic)
                if is_first_user_message_for_display:
                    with st.chat_message('user'):
                        st.write(content)
                    is_first_user_message_for_display = False
                
                # 2. Add ALL user prompts to message_history
                if content is not None:
                    history_parts_for_this_message.append(str(content))

            # --- Assistant Message Handling ---
            # 3. When role is assistant add to message_history more sparingly
            elif role == 'assistant':
                with st.chat_message('assistant'): # Streamlit display context
                    if isinstance(content, dict):
                        for item_type, item_data in content.items():
                            # --- Code Block ---
                            if item_type == 'code':
                                with st.expander("Code & Work", expanded=False):
                                    st.code(item_data, language='python')
                                # No specific instruction in this block to add 'code' type to history.

                            # --- Result Block ---
                            # Comment: "# add this to message_history IF the type is a pandas dataframe. AND if it is a pandas dataframe then input it as a str"
                            elif item_type.startswith('result') and item_data is not None:
                                utils.show_output(item_data) # Display
                                if isinstance(item_data, pd.DataFrame):
                                    history_parts_for_this_message.append(item_data.to_string())
                                else:
                                    # If it's a result but not a DataFrame, still add its string representation
                                    # to adhere to "simplify ... you simply have the assistant message".
                                    history_parts_for_this_message.append(str(item_data))

                            # --- DeepInsights Block ---
                            # Comment: "# for deepinsights only add to message_history if it is the report"
                            elif item_type.startswith('deepinsights'):
                                if isinstance(item_data, dict): # item_data for deepinsights should be a dict
                                    for subitem_name, subitem_content in item_data.items(): # Renamed 'subitem' to 'subitem_content'
                                        # Display logic for deepinsights parts
                                        # if subitem_name == 'plan':
                                        #     st.write("This task has been assigned to **deepinsights agent.**")
                                        #     st.write("**Here is the plan that I've crafted**")
                                        #     if isinstance(subitem_content, list):
                                        #         for i, subplan in enumerate(subitem_content, 1):
                                        #             st.write(f'Step {i}')
                                        #             st.write(subplan)
                                        if subitem_name == 'plan_findings':
                                            with st.expander('Task Execution and Evidence Gathering'):
                                                if isinstance(subitem_content, list):
                                                    for exec_detail in subitem_content: # Renamed 'i' loop var
                                                        if isinstance(exec_detail, (list, tuple)) and len(exec_detail) == 2:
                                                            st.code(exec_detail[0])
                                                            utils.show_output(exec_detail[1])
                                        elif subitem_name == 'report':
                                            st.write(subitem_content)
                                            # History: "# only add this portion" (the report)
                                            history_parts_for_this_message.append(str(subitem_content))
                                        elif subitem_name == 'chart_for_report':
                                            # st.write('\n')
                                            # if hasattr(subitem_content, 'savefig'): # Single Matplotlib figure
                                            #     st.plotly_chart(subitem_content)
                                            # elif isinstance(subitem_content, list): # List of figures or other data
                                            #     if not subitem_content:
                                            #         st.write("Item 'report_visual' is an empty list. Nothing to display.")
                                            #     else:
                                            #         are_all_figures = all(hasattr(item, 'savefig') for item in subitem_content)
                                            #         if are_all_figures:
                                            #             st.write(f"Displaying {len(subitem_content)} Matplotlib figure(s) from a list.")
                                            #             for fig_object in subitem_content:
                                            #                 st.plotly_chart(fig_object)
                                            #         else:
                                            #             st.write("Item 'report_visual' is a list, but not all its elements are Matplotlib figures.")
                                            #             st.write(subitem_content)
                                            # elif subitem_content is not None : # If not a figure or list, but not None
                                            #     st.write("Item 'report_visual' content:")
                                            #     st.write(subitem_content)
                                            utils.show_output(subitem_content)
                                        else: # Other subitems in deepinsights (not 'plan', 'plan execution', 'report', 'report_visual')
                                            st.write(subitem_content) # Display them
                                else: # If item_data for 'deepinsights' is not a dict
                                    st.write(item_data) # Display as is

                            # --- React Thread Block ---
                            # Comment: "# for react_thread only add to message_history IF it is the key iteration item"
                            # Comment: "# add only this part to message_history -> convert to str first. key_iteration = list(item_data.keys())[-2]"
                            elif item_type.startswith("react_thread"):
                                react_answer_for_display = None # To store what's shown under "**Answer**"
                                if isinstance(item_data, dict): # item_data for react_thread is a dict of iterations
                                    with st.expander("Solving Problem", expanded=False):
                                        for iteration, iteration_items in item_data.items():
                                            st.write(f"**Iteration {iteration}**")
                                            st.write(iteration_items.get("reason_output")) # Use .get for safer access
                                            st.code(iteration_items.get("act_work"), language="python")

                                            if iteration_items.get("reason_output", {}).get("finished") is False:
                                                st.write(iteration_items.get("act_output"))
                                    
                                    # Determine and display the "Answer" part
                                    iteration_keys = list(item_data.keys())
                                    if len(iteration_keys) >= 2:
                                        key_for_answer = iteration_keys[-1]
                                        react_answer_for_display = item_data.get(key_for_answer, {}).get("act_output")
                                    elif len(iteration_keys) == 1: # Fallback for display if only one iteration
                                        key_for_answer = iteration_keys[0]
                                        react_answer_for_display = item_data.get(key_for_answer, {}).get("act_output")
                                    # Add further fallback for display if act_output was None (e.g. check reason_output's answer field)
                                    if react_answer_for_display is None and key_for_answer and item_data.get(key_for_answer, {}).get("reason_output", {}).get("finished") is True :
                                        react_answer_for_display = item_data.get(key_for_answer, {}).get("reason_output", {}).get("answer")


                                    st.write(f"**Answer**")
                                    if react_answer_for_display is not None:
                                        st.write(react_answer_for_display)
                                    
                                    # Add to history: "only add this part... key_iteration = list(item_data.keys())[-2]"
                                    if len(iteration_keys) >= 2:
                                        key_for_history = iteration_keys[-2]
                                        act_output_for_history = item_data.get(key_for_history, {}).get("act_output")
                                        if act_output_for_history is not None:
                                            history_parts_for_this_message.append(str(act_output_for_history))
                                        else: # If act_output is None, check if reason_output has a final answer for that key_for_history
                                            reason_output_hist = item_data.get(key_for_history, {}).get("reason_output",{})
                                            if reason_output_hist.get("finished") is True and reason_output_hist.get("answer") is not None:
                                                history_parts_for_this_message.append(str(reason_output_hist.get("answer")))

                                    # No history addition from react_thread if less than 2 iterations based on the [-2] rule.
                                else: # item_data for react_thread is not a dict
                                     st.write(item_data)


                            # --- Other item_types in assistant's dict content ---
                            # Comment: "# add this to message_history" (associated with the 'else' displaying item_data)
                            else:
                                st.write(item_data) # Display
                                if item_data is not None:
                                    history_parts_for_this_message.append(str(item_data))
                    
                    # --- Assistant content is a simple string (not a dict) ---
                    else:
                        st.write(content) # Display
                        # For history: "simplify ... you simply have the assistant message"
                        if content is not None:
                            history_parts_for_this_message.append(str(content))

            # --- Consolidate and add the simplified message to history list for the current prompt_id ---
            if history_parts_for_this_message:
                # Join all parts collected for this message, filter out None/empty strings after str conversion,
                # and strip whitespace from the final string.
                final_simplified_content = "\n".join(filter(None, [str(p) if p is not None else "" for p in history_parts_for_this_message])).strip()
                if final_simplified_content: # Add only if there's non-empty content
                    current_prompt_history_list.append({
                        "role": role,
                        "content": final_simplified_content
                    })
        
        # Store the generated simplified history for this prompt_id
        # This will overwrite any previous history for this prompt_id, ensuring it's up-to-date.
        st.session_state.message_history[prompt_id] = current_prompt_history_list

def _populate_message_history_object():
    """
    Populates st.session_state.message_history with a simplified version
    based on the current st.session_state.messages. No UI output.
    This ensures the message_history object is up-to-date.
    """
    if 'messages' not in st.session_state or not st.session_state.messages:
        st.session_state.message_history = {} # Ensure it's empty if no messages
        return

    # Initialize if not present or not a dict (though above check might cover this)
    if 'message_history' not in st.session_state or not isinstance(st.session_state.message_history, dict):
        st.session_state.message_history = {}

    processed_message_history = {} # Work with a temporary dict to build fresh

    for prompt_id, messages_list in st.session_state.messages.items():
        current_prompt_history_list = []
        for message_obj in messages_list: # Renamed 'message' to 'message_obj'
            role = message_obj.get('role', 'unknown')
            content = message_obj.get('content')
            history_parts_for_this_message = []

            if role == 'user':
                if content is not None:
                    history_parts_for_this_message.append(str(content))
            elif role == 'assistant':
                if isinstance(content, dict):
                    for item_type, item_data in content.items():
                        # This logic should mirror exactly what display_chat_history
                        # considers for its simplified history string parts.
                        if item_type.startswith('result') and item_data is not None:
                            if isinstance(item_data, pd.DataFrame):
                                history_parts_for_this_message.append(item_data.to_string())
                            else:
                                history_parts_for_this_message.append(str(item_data))
                        elif item_type.startswith('deepinsights'):
                            if isinstance(item_data, dict):
                                report_content = item_data.get('report')
                                if report_content is not None:
                                    history_parts_for_this_message.append(str(report_content))
                        elif item_type.startswith("react_thread"):
                            if isinstance(item_data, dict):
                                iteration_keys = list(item_data.keys())
                                key_for_history = None
                                if len(iteration_keys) >= 2:
                                    key_for_history = iteration_keys[-2]
                                elif len(iteration_keys) == 1:
                                    key_for_history = iteration_keys[0]
                                
                                if key_for_history:
                                    act_output_hist = item_data.get(key_for_history, {}).get("act_output")
                                    if act_output_hist is not None:
                                        history_parts_for_this_message.append(str(act_output_hist))
                                    else:
                                        reason_output_hist = item_data.get(key_for_history, {}).get("reason_output", {})
                                        if reason_output_hist.get("finished") is True and reason_output_hist.get("answer") is not None:
                                            history_parts_for_this_message.append(str(reason_output_hist.get("answer")))
                        elif item_type != 'code' and item_type != 'report_visual' and item_type != 'plan' and item_type != 'plan execution': 
                            # Avoid adding raw code, visuals, or plans unless explicitly part of a "result" or "report"
                            # This 'else' covers other item_types that should be stringified as per your original logic.
                            if item_data is not None:
                                history_parts_for_this_message.append(str(item_data))
                elif content is not None: # Assistant content is a simple string
                    history_parts_for_this_message.append(str(content))
            
            if history_parts_for_this_message:
                final_simplified_content = "\n".join(filter(None, [str(p).strip() if p is not None else "" for p in history_parts_for_this_message])).strip()
                if final_simplified_content:
                    current_prompt_history_list.append({
                        "role": role,
                        "content": final_simplified_content
                    })
        processed_message_history[prompt_id] = current_prompt_history_list
    
    st.session_state.message_history = processed_message_history


def load_daily_log(filepath=DAILY_LOG_FILEPATH):
    """Loads the daily log from a JSON file. Returns empty dict if file not found or invalid JSON."""
    if not os.path.exists(filepath):
        return {}
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            # vvv--- MODIFIED LINES START ---vvv
            content = f.read()
            if not content: # Handle empty file
                return {}
            return json.loads(content)
            # ^^^--- MODIFIED LINES END ---^^^
    except json.JSONDecodeError: # More specific error catch
        return {} 
    
def save_daily_log(data, filepath=DAILY_LOG_FILEPATH):
    """Saves the daily log data to a JSON file."""
    try:
        # vvv--- ADDED BLOCK START ---vvv
        directory = os.path.dirname(filepath)
        if directory and not os.path.exists(directory): # Check if directory string is not empty
            os.makedirs(directory)
        # ^^^--- ADDED BLOCK END ---^^^
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
    except Exception as e:
        st.sidebar.error(f"Error saving daily log: {e}")

def update_and_log_daily_history():
    """
    Updates the daily log by appending the current st.session_state.message_history
    to a list of sessions for the current day.
    """
    if 'message_history' not in st.session_state or not st.session_state.message_history:
        # st.sidebar.warning("No message history in current session to log.")
        return

    month_name, week_number, date_day_key = utils.get_current_time_components()
    daily_log_data = load_daily_log()

    month_entry = daily_log_data.setdefault(month_name, {})
    week_container = month_entry.setdefault("week", {})
    week_entry = week_container.setdefault(str(week_number), {}) # Ensure week_number is string key
    date_container = week_entry.setdefault("date", {})
    
    day_sessions_list = date_container.get(date_day_key)
    if not isinstance(day_sessions_list, list):
        day_sessions_list = []
        
    if st.session_state.message_history: # Ensure there's something to copy and append
        current_session_history_copy = copy.deepcopy(st.session_state.message_history)
        day_sessions_list.append(current_session_history_copy)
    
    # Only update the date_container if there's actual data to store
    if day_sessions_list: 
        date_container[date_day_key] = day_sessions_list
    # Optional: if day_sessions_list is empty but the key existed, you might want to remove the key
    # or ensure it's an empty list. Current logic stores an empty list if it was made empty.

    save_daily_log(daily_log_data)
    print(f"Chat history for {date_day_key} processed for logging.")


def data_analyst_tab(dataframes_dict, user_input=None, equations_dict=None, pdf_file_obj=None): # Accept user_input as argument
    """Encapsulates the Data Analyst chat functionality."""

    # Initialize session state for messages if it doesn't exist
    if 'messages' not in st.session_state:
        st.session_state.messages = {}
       
    if 'prompt_id' not in st.session_state:
        st.session_state.prompt_id = 0

    #display_chat_history() # Display history first

    # --- Input processing is now handled if user_input is passed from main_app ---
    if user_input:
        current_prompt_id = st.session_state.prompt_id
        if current_prompt_id not in st.session_state.messages:
            st.session_state.messages[current_prompt_id] = []

        # Add user message to the history for the current prompt
        # Check if the last message for this prompt_id is already this user input to avoid duplication on reruns
        if not st.session_state.messages[current_prompt_id] or st.session_state.messages[current_prompt_id][-1].get('content') != user_input:
            st.session_state.messages[current_prompt_id].append({"role": "user", "content": user_input})

            # Prepare local variables for the agent
            local_var = {'dataframes_dict': dataframes_dict}
            if len(pdf_file_obj) > 0:
                local_var["pdf_file_objects"] = pdf_file_obj

            with st.chat_message("assistant"):
                with st.spinner("Thinking...", show_time=True):
                    try:
                        user_instructions = load_instructions()
                        user_input = f"User prompt: {user_input}"
                        if user_instructions is not None:
                            user_input += f"\n\nGeneral user instructions (follow carefully **if** applicable. May not apply to you, this is a multi-agent system.): {user_instructions}"
                        if equations_dict is not None:
                            user_input += f"\n\nHere are the list of equations, follow accordingly to calculate various features and metrics: {equations_dict}"

                        agent = DataAnalystAgent(user_input, local_var)
                        # Assuming agent.main() adds results to session state via utils.assistant_message
                        agent.main()

                    except Exception as e:
                        tb_str = traceback.format_exc()

                        # 2. Display a more informative error in Streamlit UI
                        #    Include the exception type and message. Mention full details are logged.
                        st.error(f"An error occurred during analysis: {type(e).__name__}: {e}. Check console log for details.")

                        # 3. Create error content for session state (optional: add more detail if desired)
                        #    Keeping it relatively simple for the UI message store.
                        error_detail = f"Error Type: {type(e).__name__}\nMessage: {e}\n(Full traceback logged to console)"
                        
                        error_content = {"error": error_detail}

                        # 4. Append error content to Streamlit session state (Ensure this logic is correct for your app)
                        #    Make sure 'current_prompt_id' is defined in the scope where this code runs.
                        try:
                            if 'messages' in st.session_state and current_prompt_id in st.session_state.messages:
                                st.session_state.messages[current_prompt_id].append({"role": "assistant", "content": error_content})
                            else:
                                # Handle cases where session state might not be initialized as expected
                                # Maybe add the error to a general log or display differently
                                print(f"Warning: Could not append error to st.session_state.messages for prompt ID {current_prompt_id}")
                                # Fallback: maybe display the error_detail directly if session state fails?
                                # st.json(error_content) # Example fallback
                        except NameError:
                            print("Warning: 'current_prompt_id' not defined in this scope. Cannot save error to session state.")
                            # Handle error display without session state if needed


                        # 5. Print the FULL traceback to the console (or ideally, a log file)
                        #    This gives YOU the exact file, line number, and call stack.
                        print("--- ERROR TRACEBACK ---")
                        print(tb_str)
                        print("-----------------------")
                        # You might also want to log the simpler error_content that was shown/saved
                        # print("Error content for UI:", error_content)

    # --- END MODIFIED ERROR HANDLING ---
            # Increment prompt_id for the *next* interaction
            st.session_state.prompt_id += 1
            # Rerun needed to display the latest user message and assistant response correctly
            #st.rerun()

def instructions_tab():
    """Encapsulates the User Instructions tab functionality."""
    st.header("User Instructions")
    st.write("Define custom instructions or context for the AI agents below.")
    st.caption("These instructions will be saved and loaded automatically.")

    current_instructions = load_instructions()
    instructions_key = "user_instructions_text_area"
    updated_instructions = st.text_area(
        "Instructions",
        value=current_instructions,
        height=300,
        key=instructions_key
    )

    if st.button("Save Instructions", key="save_instructions_button"):
        if save_instructions(updated_instructions):
            st.success("Instructions saved successfully!")
            st.rerun()
        else:
            st.error("Failed to save instructions.")


def feedback_tab():
    st.header("Feedback")
    st.write("Provide below what you think the agent could get better at. Your feedback will be emailed directly to the Data Science team.")
    st.write('Leave your name ')
    # Initialize session state variables
    if "show_success" not in st.session_state:
        st.session_state["show_success"] = False
    if "clear_feedback" not in st.session_state:
        st.session_state["clear_feedback"] = False
    
    # Clear the text area if the clear_feedback flag is set
    if st.session_state["clear_feedback"]:
        st.session_state["user_feedback_text_area_widget"] = ""
        st.session_state["clear_feedback"] = False
    
    # Create the text area widget
    feedback_text = st.text_area(
        "User Feedback Window",
        height=300,
        key="user_feedback_text_area_widget"
    )
    
    # Handle feedback submission
    if st.button("Send Feedback", key="save_feedback_button"):
        if feedback_text:
            try:
                # Save feedback with timestamp
                now = datetime.datetime.now()
                timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
                feedback_entry = (
                    f"Timestamp: {timestamp}\n"
                    f"Feedback: \n{feedback_text}\n"
                    f"{'-' * 40}\n\n\n\n"
                )
                with open(FEEDBACK_FILE, 'a', encoding='utf-8') as f:
                    f.write(feedback_entry)

                # send feedback as email so every time you get feedback you get the list. 
                with open(FEEDBACK_FILE, 'r') as f:
                    body = f.read()
                    email = 'yasser.x.ali@kp.org'
                    subject = 'Data Analyst Agent Feedback'
                    send_email(email, subject, body)
                    

                # Set flags to clear text area and show success message
                st.session_state["clear_feedback"] = True
                
                st.session_state["show_success"] = True
                st.rerun()
            except Exception as e:
                st.error(f"Error saving feedback: {e}")
        else:
            st.warning("Please enter your feedback before trying to save.")
    
    # Show success message temporarily if flag is set
    if st.session_state["show_success"]:
        st.success("Feedback sent successfully! Thank you very much :) ")
        time.sleep(3)  # Display message for 4 seconds
        st.session_state["show_success"] = False
        st.rerun()

def convert_data_for_json(data):
    """
    Recursively converts objects not directly serializable by JSON.
    Specifically handles Pandas DataFrames and Series.
    """
    if isinstance(data, list):
        return [convert_data_for_json(item) for item in data]
    elif isinstance(data, dict):
        new_dict = {}
        for key, value in data.items():
            new_dict[key] = convert_data_for_json(value)
        return new_dict
    elif isinstance(data, pd.DataFrame):
        # Convert DataFrame to a list of dictionaries (records orientation)
        # Other orientations: 'split', 'index', 'columns', 'values'
        # Choose the one that best suits how you want to reconstruct it later.
        return data.to_dict(orient='records')
    elif isinstance(data, pd.Series):
        # Convert Series to a list
        return data.to_list()
    # Add any other custom type conversions here if needed
    # For example, converting datetime objects to ISO format strings:
    # elif isinstance(data, datetime.datetime):
    #     return data.isoformat()
    else:
        return data


# --- Main App ---
def main_app():
    st.set_page_config(layout="wide")
    st.title('Data Analyst Agent')
    
    if 'dataframes_dict' not in st.session_state:
        st.session_state.dataframes_dict = {}

    if 'messages' not in st.session_state:
        st.session_state.messages = {}

    if 'prompt_id' not in st.session_state:
        st.session_state.prompt_id = 0

    if 'message_history' not in st.session_state:
        st.session_state.message_history = {}

    if 'processed_sheets_cache' not in st.session_state:
        st.session_state.processed_sheets_cache = {}

    if 'pdf_file_objects' not in st.session_state:
        st.session_state.pdf_file_objects = {}

    # --- 2. ESSENTIAL: Corrected Sidebar Logic ---
    with st.sidebar:
        st.header("Data Upload / Management")
        newly_uploaded_files_dict, equations_dict = upload_and_format_files()


        # Corrected logic to MERGE uploaded files into st.session_state.dataframes_dict
        if newly_uploaded_files_dict:
            for key, new_data_item in newly_uploaded_files_dict.items():
                st.session_state.dataframes_dict[key] = new_data_item

    
        utils.sync_file_metadata_from_session()

        # Display all available data sources (from SQL and all uploads)
        # probably turn this into a function and move it to utils_for_main.py for easier readability. This file is wayyyyy to long. 
        st.subheader("Available Data Sources:")
        if st.session_state.dataframes_dict:
            for name, item in sorted(list(st.session_state.dataframes_dict.items())):
                with st.expander(name, expanded=False):
                    if isinstance(item, pd.DataFrame):
                        st.write(item)
                    elif isinstance(item, dict): 
                        # Check if it's PDF dict (int keys) or Excel (str keys)
                        if all(isinstance(key, int) for key in item.keys()):  # PDF output
                            for page_num in sorted(item.keys()):
                                page = item[page_num]
                                with st.expander(f"Page {page_num}", expanded=False):
                                    # Display text
                                    st.subheader("Extracted Text:")
                                    if isinstance(page["text"], list):
                                        for text_block in page["text"]:
                                            if isinstance(text_block, dict) and "text" in text_block:
                                                st.write(text_block["text"])
                                            else:
                                                st.write(text_block)  # Fallback
                                    else:
                                        st.write(page["text"])

                                    # Display tables
                                    st.subheader("Extracted Tables:")
                                    for idx, table in enumerate(page.get("tables", []), start=1):
                                        st.write(f"Table {idx}:")
                                        cleaned_data = table.get("cleaned_table")
                                        if cleaned_data and isinstance(cleaned_data, list):
                                            try:
                                                df_table = pd.DataFrame(cleaned_data)
                                                st.write(df_table)
                                            except Exception as e:
                                                st.write("Could not display table as DataFrame:", str(e))
                                                st.write(cleaned_data)  # Fallback to raw
                                        else:
                                            st.write(table.get("raw_data", "No table data"))
                        else:  # Assume Excel or other dict with str keys
                            for sheet_name, sheet in sorted(list(item.items())):
                                st.write(f"Sheet: {sheet_name}") # sheet_name_display is already capitalized
                                st.write(sheet)
                    else:
                        st.write(f"- {name} (Unknown type)")
        else:
            st.warning("No data files loaded yet. Upload files or use the SQL agent.")

    tab1, tab2, tab3 = st.tabs(["Data Analyst", "User Instructions", "User Feedback"])

    with tab1:
        display_chat_history()
    with tab2:
        instructions_tab()
    with tab3:
        feedback_tab()



    user_input = st.chat_input('What can I help with?', key='main_chat_input') # Your original key

    if user_input:
        with st.chat_message('user'):
            st.write(user_input)
        
        data_analyst_tab(st.session_state.dataframes_dict, user_input, equations_dict=equations_dict, pdf_file_obj=st.session_state.pdf_file_objects)

        _populate_message_history_object()
        update_and_log_daily_history()
        
      
        st.rerun()
        

if __name__ == '__main__':
    main_app()