import streamlit as st
import utils 
import os
import pandas as pd
import numpy as np
import re 
from agents import Agent
import openpyxl
import math
import io
import time 



def _make_column_names_unique(column_names):
    name_counts = {}
    new_columns = []
    for col in column_names:
        if col is None:
            col_str = "Unnamed"
        elif not isinstance(col, str):
            col_str = str(col).strip() # Good: handles non-strings, strips whitespace
        else:
            col_str = col.strip()      # Good: handles strings, strips whitespace

        if not col_str: 
            col_str = "Unnamed"        # Good: handles empty strings after stripping
        
        # Current logic:
        if col_str not in name_counts:
            name_counts[col_str] = 1
            new_columns.append(col_str) 
        else:
            name_counts[col_str] += 1
            new_columns.append(f"{col_str}_{name_counts[col_str]}")
    return new_columns

def load_data(uploaded_file_obj):
    """
    Loads data from an uploaded file.
    Returns a pandas DataFrame for CSVs or a dictionary of DataFrames for Excel files.
    """
    try:
        uploaded_file_obj.seek(0)
        file_name = uploaded_file_obj.name.lower()
        if file_name.endswith('.csv'):
            return pd.read_csv(uploaded_file_obj)
        elif file_name.endswith(('.xls', '.xlsx', '.xlsm', '.xlsb')):
            # Pass the file object directly, openpyxl (used by read_excel) can handle it
            return pd.read_excel(uploaded_file_obj, sheet_name=None, engine='openpyxl')
        else:
            st.error(f"Unsupported file type: {uploaded_file_obj.name}")
            return None
    except Exception as e:
        st.error(f'Error reading file {uploaded_file_obj.name}: {e}')
        return None


def _process_dataframe(df):
    """
    A helper function to apply a standard set of cleaning and transposing rules.
    This version also removes columns that are duplicates by content.
    """
    # 1. Remove rows and columns that are completely empty
    cleaned_df = df.dropna(how='all', axis=0).dropna(how='all', axis=1)

    if cleaned_df.empty:
        return cleaned_df
        

    # 3. Set the first column as the index to prepare for transposing
    feature_col_name = cleaned_df.columns[0]
    feature_col_series = cleaned_df[feature_col_name].fillna('Unnamed Feature') # Ensure index has no NaNs
    indexed_df = cleaned_df.set_index(feature_col_series)
    if feature_col_name in indexed_df.columns: # Drop the original column if it wasn't the only one
         indexed_df = indexed_df.drop(columns=feature_col_name)
    
    # 4. Transpose the DataFrame
    transposed_df = indexed_df.T

    # 5. Make column names unique (handling name collisions)
    if not transposed_df.empty:
        transposed_df.columns = _make_column_names_unique(list(transposed_df.columns))
    
    # --- NEW STEP 6: Remove columns that are duplicates by content ---
    if not transposed_df.empty:
        columns_to_keep = []
        seen_column_signatures = set()

        for col_name in transposed_df.columns:
            current_series = transposed_df[col_name]
            # Create a signature: convert to string, replace 'nan', make tuple
            # This makes NaNs comparable for uniqueness
            series_signature = tuple(current_series.astype(str).str.replace('nan', '_NaN_Placeholder_').values)
            
            if series_signature not in seen_column_signatures:
                columns_to_keep.append(col_name)
                seen_column_signatures.add(series_signature)
            # Else, this column's content is a duplicate of one already kept, so we skip it.
        
        transposed_df = transposed_df[columns_to_keep]

    return transposed_df



def convert_sum_to_addition(formula):
    """Converts simple SUM formulas like '=SUM(A1,B2)' to '=A1+B2'."""
    match = re.match(r'=SUM\(([^:]+)\)', formula, re.IGNORECASE)
    if match and ',' in match.group(1):
        args = [arg.strip() for arg in match.group(1).split(',')]
        return '=' + '+'.join(args)
    return formula


def extract_equations_from_sheet(file_source, sheet_name=None) -> dict:
    """
    Analyzes an Excel sheet and returns ONLY a dictionary of human-readable
    vertical equations, mapping feature names to their formulas.
    """
    try:
        # Load workbook for both formulas and their calculated values
        if isinstance(file_source, str):
            wb_formulas = openpyxl.load_workbook(file_source, data_only=False)
            wb_values = openpyxl.load_workbook(file_source, data_only=True)
        else:
            file_source.seek(0)
            file_content_bytes = file_source.read()
            wb_formulas = openpyxl.load_workbook(io.BytesIO(file_content_bytes), data_only=False)
            wb_values = openpyxl.load_workbook(io.BytesIO(file_content_bytes), data_only=True)
        
        # Select the correct sheet
        if sheet_name and sheet_name in wb_formulas.sheetnames:
            sheet_formulas, sheet_values = wb_formulas[sheet_name], wb_values[sheet_name]
        else:
            sheet_formulas, sheet_values = wb_formulas.active, wb_values.active
    except Exception as e:
        # Use a more generic error message or logging
        print(f"Error loading workbook: {e}")
        return {}

    # 1. Create a raw DataFrame from values to map feature names
    initial_data = [[cell.value for cell in row] for row in sheet_values.iter_rows()]
    df_raw = pd.DataFrame(initial_data)

    # 2. Generate a unique name for each potential feature row
    temp_df = df_raw.dropna(how='all', axis=0).dropna(how='all', axis=1)
    if temp_df.empty:
        return {}
    
    min_non_empty = math.ceil(len(temp_df.columns) * 0.6)
    temp_df = temp_df.dropna(thresh=min_non_empty, axis=0)

    if temp_df.empty:
        return {}
        
    feature_col_series = temp_df[temp_df.columns[0]].fillna('Unnamed Feature')
    unique_names = _make_column_names_unique(list(feature_col_series))
    row_index_to_unique_name = dict(zip(temp_df.index, unique_names))

    # 3. Build the coordinate-to-unique-name map for formula translation
    coord_to_unique_name_map = {}
    for row_idx, unique_name in row_index_to_unique_name.items():
        for col_idx in range(sheet_formulas.max_column):
            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
            coord = f"{col_letter}{row_idx + 1}"
            coord_to_unique_name_map[coord] = unique_name

    # 4. Iterate through the sheet to find and translate formulas
    equations_dict = {}
    for row in sheet_formulas.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                formula_string = cell.value
                # Exclude formulas linking to other sheets or files
                if '!' in formula_string or '[' in formula_string:
                    continue

                # Check if the formula is vertical (references cells in other rows)
                is_horizontal = True
                referenced_rows = [int(r) for r in re.findall(r'[A-Z]+([0-9]+)', formula_string)]
                if any(ref_row != cell.row for ref_row in referenced_rows):
                    is_horizontal = False
                
                if not is_horizontal:
                    key = coord_to_unique_name_map.get(cell.coordinate)
                    if key and key not in equations_dict:
                        processed_formula = convert_sum_to_addition(formula_string)
                        readable_formula = processed_formula.lstrip('=')
                        
                        # Replace cell coordinates with unique feature names
                        cell_refs = re.findall(r'([A-Z]+[0-9]+)', readable_formula)
                        for ref in sorted(list(set(cell_refs)), key=len, reverse=True):
                            mapped_name = coord_to_unique_name_map.get(ref, ref)
                            readable_formula = re.sub(r'\b' + ref + r'\b', f"'{mapped_name}'", readable_formula)

                        # Clean up spacing for readability
                        readable_formula = re.sub(r'([*\/+\-])', r' \1 ', readable_formula).strip()
                        readable_formula = re.sub(r'\s+', ' ', readable_formula)
                        
                        equations_dict[key] = readable_formula
                        
    return equations_dict



def create_compact_summary(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Creates a highly condensed and annotated "skeleton" view of a DataFrame for an LLM.

    This function aggressively compacts the data by:
    1. Ignoring all completely empty rows.
    2. For each row, gathering only the non-empty cells.
    3. If a row has more than 4 non-empty cells, it keeps only the first 2
       and last 2, placing '...' in between.
    4. Annotating each kept cell with its original (row, col) coordinate.
    """
    all_processed_rows = []

    # Iterate through each row of the original raw DataFrame
    for r_idx, row in df_raw.iterrows():
        # 1. Gather only the non-empty cells from the row, keeping their original coordinates
        content_items = []
        for c_idx, cell_value in enumerate(row):
            if pd.notna(cell_value) and str(cell_value).strip() != '':
                content_items.append({'value': cell_value, 'r': r_idx, 'c': c_idx})

        # If the row was effectively empty, skip it entirely
        if not content_items:
            continue

        # 2. Condense the list of content items if it's too long
        condensed_items = []
        if len(content_items) > 4:
            # Keep first two, add ellipsis, keep last two
            condensed_items.extend(content_items[:2])
            condensed_items.append({'value': '...', 'r': None, 'c': None}) # Ellipsis marker
            condensed_items.extend(content_items[-2:])
        else:
            # If the row is short, keep all its content
            condensed_items = content_items

        # 3. Format each item for the final output row
        formatted_row = []
        for item in condensed_items:
            if item['value'] == '...':
                formatted_row.append('...')
            else:
                # Annotate with original coordinates
                formatted_row.append(f"{item['value']} ({item['r']}, {item['c']})")
        
        all_processed_rows.append(formatted_row)

    # 4. Create a final, compact DataFrame from the processed rows
    # Pandas will automatically handle the jagged rows by filling with None/NaN
    summary_df = pd.DataFrame(all_processed_rows)

    # Set clean, generic column headers
    summary_df.columns = [f'Col_{i}' for i in range(summary_df.shape[1])]

    return summary_df




def convert_to_float_if_numeric(df: pd.DataFrame) -> pd.DataFrame:
    """
    Goes column by column through a DataFrame and converts the
    data type to float if the entire column is numeric.

    Args:
        df: The input pandas DataFrame.

    Returns:
        The DataFrame with numeric columns converted to float type.
    """
    for col in df.columns:
        # Attempt to convert the column to a numeric type.
        # 'errors=coerce' will turn any non-numeric values into NaN (Not a Number).
        numeric_series = pd.to_numeric(df[col], errors='coerce')

        # If the conversion doesn't result in all values being NaN,
        # it means the column was numeric. Then we can change the type.
        if not numeric_series.isnull().all():
            df[col] = numeric_series.astype(float)
    return df


def get_sections_agent(compact_df):
    """
    The purpose of this function is to create an agent that can extract the section names from the compact dataframe function above
    """

    agent = Agent()

    system_prompt = """
    You are a top-tier document analysis agent. Your specialty is analyzing structured text representations of spreadsheets to identify their semantic layout.

    **CONTEXT OF THE INPUT:**
    You will be given a string representation of a pandas DataFrame. This DataFrame is a 'compact summary' of a larger, more complex spreadsheet. This summary has several key features you MUST understand:
    1.  **Coordinate Annotations:** Each cell contains the original value followed by its zero-based `(row, column)` coordinate from the source document, like `Q1 Revenue & Costs (2, 1)`.
    2.  **Content Condensing:** Rows with many values have been condensed. They show only the first two and last two non-empty items, with `...` placed in between.
    3.  **Generic Headers:** The column headers of the summary are generic (e.g., `Col_0`, `Col_1`, `Col_2`). You should ignore them and focus on the content within the cells.

    **YOUR TASK:**
    Your sole task is to identify the titles of the major data sections within this summary and return them as a list.
    - A section title is a prominent piece of text, usually in the first content column, that marks the beginning of a new logical block of data.
    - You must list the titles in the order they appear in the document from top to bottom.
    - Do not include the main document title (usually on row 0), only the titles of the subsections.

    **OUTPUT REQUIREMENTS:**
    - Your output MUST be a single, valid JSON object and nothing else.
    - The JSON object must have a single key named "section_titles_list".
    - The value for this key must be a JSON array (a list) of strings, where each string is a section title you identified.

    ---
    **EXAMPLE OF REQUIRED TRANSFORMATION:**

    **INPUT TEXT (This is the format you will receive):**  

                                Col_0                        Col_1                             Col_2                         Col_3
    0           Consolidated P&amp;L (0, 1)
    1         Q1 Revenue & Costs (2, 1)
    2                 Revenue (3, 1)              15000.0 (3, 2)              15250.0 (3, 3)              ...              15500.0 (3, 5)
    3                    COGS (4, 1)               (6000.0) (4, 2)               (6100.0) (4, 3)              ...               (6200.0) (4, 5)
    4
    5      Operating Expenses (6, 1)
    6               Marketing (7, 1)                 (500.0) (7, 2)                (520.0) (7, 3)
    7                    R&amp;D (8, 1)                (1200.0) (8, 2)               (1250.0) (8, 3)
    8
    9        Headcount Summary (10, 1)
    10             Engineers (11, 1)                     45 (11, 2)
    11                 Sales (12, 1)                     15 (12, 2)


    **REQUIRED JSON OUTPUT:**
    ```json
    {
    "section_titles_list": [
        "Q1 Revenue & Costs",
        "Operating Expenses",
        "Headcount Summary"
    ]
    }

    """

    user_input = f"""
    Please analyze the following spreadsheet summary and return the structured JSON object containing the list of section titles.

    --- SPREADSHEET SUMMARY ---
    {compact_df.to_string()}
    --- END OF SUMMARY ---

    """

    json_output = agent.json_agent(system_prompt, user_input=user_input, is_visible=True)

    return json_output["section_titles_list"]


def slice_dataframe_by_sections(df_raw: pd.DataFrame, section_titles: list) -> dict:
    """
    Slices a raw DataFrame into a dictionary of cleaned, separate DataFrames
    by performing a full search for each section title to find its exact location.

    This version is based on the robust "search all cells" strategy and ensures
    all resulting DataFrames have unique column names.

    Args:
        df_raw (pd.DataFrame): The original, untouched DataFrame from the Excel sheet.
        section_titles (list): A list of strings corresponding to the section titles.

    Returns:
        dict: A dictionary where keys are snake_cased section titles and
              values are the corresponding cleaned pandas DataFrames.
    """
    final_dataframes = {}
    
    if df_raw.empty:
        return {}

    # --- 1. Search every cell to find the exact (row, col) of each title ---
    title_locations = {}
    title_set = set(section_titles)
    
    print("--- Searching for section titles in the DataFrame ---")
    for r_idx in range(len(df_raw)):
        for c_idx in range(len(df_raw.columns)):
            cell_value = str(df_raw.iloc[r_idx, c_idx]).strip()
            if cell_value in title_set:
                # Store the title and its precise (row, col) coordinate
                title_locations[cell_value] = (r_idx, c_idx)
                print(f"Found title '{cell_value}' at location ({r_idx}, {c_idx})")
                # Remove from set to avoid finding the same title twice
                title_set.remove(cell_value)

    if not title_locations:
        print("ERROR: None of the provided section titles were found anywhere in the DataFrame.")
        return {}
        
    print("--- Title search complete ---")

    # 2. Sort the found titles by their row number to process them in order
    sorted_titles = sorted(title_locations.items(), key=lambda item: item[1][0])

    # 3. Iterate through the sorted titles to define boundaries and slice
    for i, (title, (start_row, _)) in enumerate(sorted_titles):
        
        # Determine the end row for the current section's slice
        if i + 1 < len(sorted_titles):
            end_row = sorted_titles[i + 1][1][0] - 1 # End before the next title starts
        else:
            end_row = len(df_raw) - 1 # It's the last section, go to the end
            
        print(f"Slicing section '{title}' from row {start_row} to {end_row}...")

        # 4. Slice the original DataFrame to get the section block
        section_df = df_raw.iloc[start_row:end_row + 1].copy()

        # 5. Clean and organize the extracted block
        section_df.dropna(how='all', axis=0, inplace=True)
        section_df.dropna(how='all', axis=1, inplace=True)
        section_df.reset_index(drop=True, inplace=True)
        
        # Heuristic to find and promote the best header row within the slice
        if not section_df.empty:
            header_row_index = 0 # Default to the first row (the title row)
            min_nulls = float('inf')
            # Look for a row with the least number of empty cells to be the header
            for idx, row in section_df.iterrows():
                # Don't consider the title row itself as a potential data header if other options exist
                if idx > 0 and row.isnull().sum() < min_nulls:
                    min_nulls = row.isnull().sum()
                    header_row_index = idx
            
            # Set the new headers and drop that row from the data
            if len(section_df) > header_row_index:
                # *** FIX APPLIED HERE ***
                # First, assign the new headers from the identified row
                new_headers = section_df.iloc[header_row_index]
                # Then, immediately sanitize them to ensure they are unique
                section_df.columns = _make_column_names_unique(new_headers)
                
                # Finally, drop the old header row and reset the index
                section_df = section_df.drop(header_row_index).reset_index(drop=True)

        # 6. Create a clean key and add the final DataFrame to our dictionary
        key_name = re.sub(r'[^A-Z0-9_]+', '', title.replace(' ', '_').upper())
        final_dataframes[key_name] = section_df



    return final_dataframes
