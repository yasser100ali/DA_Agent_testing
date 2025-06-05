import pandas as pd
import openpyxl
import numpy as np
import math

def _make_column_names_unique(column_names):
    """
    Takes a list of column names and makes them unique by appending suffixes.
    e.g., ['A', 'B', 'A'] -> ['A', 'B', 'A.1']
    """
    seen = {}
    new_columns = []
    for col in column_names:
        # Ensure col is a string, handle None or other types gracefully
        if col is None:
            col_str = "Unnamed"
        elif not isinstance(col, str):
            col_str = str(col)
        else:
            col_str = col

        if col_str not in seen:
            seen[col_str] = 0
            new_columns.append(col_str)
        else:
            seen[col_str] += 1
            new_columns.append(f"{col_str}.{seen[col_str]}")
    return new_columns

def _process_dataframe(df):
    """
    A helper function to apply a standard set of cleaning and transposing rules.
    This version includes making column names unique.
    """
    # 1. Remove rows and columns that are completely empty
    cleaned_df = df.dropna(how='all', axis=0).dropna(how='all', axis=1)

    # 2. Remove rows that are more than 40% empty
    if cleaned_df.empty:
        return cleaned_df
        
    min_non_empty_values = math.ceil(len(cleaned_df.columns) * 0.6)
    cleaned_df = cleaned_df.dropna(thresh=min_non_empty_values, axis=0)
    
    if cleaned_df.empty:
        return cleaned_df

    # 3. Set the first column as the index to prepare for transposing
    feature_col_name = cleaned_df.columns[0]
    feature_col_series = cleaned_df[feature_col_name].fillna('Unnamed Feature')
    indexed_df = cleaned_df.set_index(feature_col_series)
    indexed_df = indexed_df.drop(columns=feature_col_name)
    
    # 4. Transpose the DataFrame
    transposed_df = indexed_df.T

    # --- NEW: Make column names unique ---
    if not transposed_df.empty:
        transposed_df.columns = _make_column_names_unique(list(transposed_df.columns))

    return transposed_df


def get_final_data_views(filepath, sheet_name=None):
    """
    Analyzes an Excel sheet and returns two cleaned, transposed DataFrames:
    1. A DataFrame with only the final calculated values.
    2. A "hybrid" DataFrame showing values and internal formulas.
    """
    try:
        wb_formulas = openpyxl.load_workbook(filepath, data_only=False)
        wb_values = openpyxl.load_workbook(filepath, data_only=True)
        
        if sheet_name and sheet_name in wb_formulas.sheetnames:
            sheet_formulas = wb_formulas[sheet_name]
            sheet_values = wb_values[sheet_name]
        else:
            sheet_formulas = wb_formulas.active
            sheet_values = wb_values.active
            
    except FileNotFoundError:
        print(f"Error: The file '{filepath}' was not found.")
        return None, None

    values_data = []
    hybrid_data = []
    for r_idx, row_cells in enumerate(sheet_formulas.iter_rows()):
        values_row = []
        hybrid_row = []
        for c_idx, cell_formulas in enumerate(row_cells):
            value = sheet_values.cell(row=r_idx + 1, column=c_idx + 1).value
            values_row.append(value)

            hybrid_content = None
            if cell_formulas.data_type == 'f':
                formula_string = cell_formulas.value
                if '!' not in formula_string and '[' not in formula_string:
                    hybrid_content = formula_string
                else:
                    hybrid_content = value
            else:
                hybrid_content = cell_formulas.value
            hybrid_row.append(hybrid_content)

        values_data.append(values_row)
        hybrid_data.append(hybrid_row)
        
    df_values_raw = pd.DataFrame(values_data)
    df_hybrid_raw = pd.DataFrame(hybrid_data)

    final_values_df = _process_dataframe(df_values_raw.copy())
    final_hybrid_df = _process_dataframe(df_hybrid_raw.copy())

    return final_values_df, final_hybrid_df

# --- USAGE ---
file_path = "testing_data/DOE_Data.xlsx"
sheet_name = "SSF_DOE"

# Assuming this script is run with Streamlit, e.g., `streamlit run your_script_name.py`

values_df_final, hybrid_df_final = get_final_data_views(file_path, sheet_name=sheet_name)


def name_unnamed_features(df):
    feature_names = list(df.columns)
    unnamed_features = {}

    for index, feature in enumerate(feature_names):
        if "UNNAMED" in feature.upper():
            unnamed_features[feature.upper()] = {
                "feature_name": feature.upper(),
                "index": index,
                "column_contents": df[feature_names[index]] 
            }

    return unnamed_features

features_info = name_unnamed_features(values_df_final)

print(features_info)